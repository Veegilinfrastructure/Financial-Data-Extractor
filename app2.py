import streamlit as st
import fitz  # PyMuPDF
import pandas as pd
import spacy
import re
from openpyxl import load_workbook
import os

# Ensure spaCy model is installed
try:
    nlp = spacy.load("en_core_web_sm")
except OSError:
    os.system("python -m spacy download en_core_web_sm")
    nlp = spacy.load("en_core_web_sm")

def extract_text_from_pdf(uploaded_file):
    """Extract text from an uploaded PDF document."""
    doc = fitz.open(stream=uploaded_file.read(), filetype="pdf")
    text = ""
    for page in doc:
        text += page.get_text()
    return text

def extract_data(text):
    """Extract key financial metrics using NLP and regex."""
    doc = nlp(text)
    data = {
        "Company Name": "",
        "Revenue": "",
        "EBITDA": "",
        "Industry": "",
        "Location": "",
        "Date": "",
        "Gross Revenues": "",
        "Owner salary": "",
        "Adjusted EBITDA": "",
        "COGS": "",
        "Gross Profit": "",
        "Total Operating Expenses": "",
        "Operating Income": "",
        "Net Pre-Tax Income": "",
        "Employees": "",
        "Interest": "",
        "Current Assets": "",
        "Cash and Cash Equivalents": "",
        "Investment": "",
        "Bank Loan": "",
        "Seller's Note Value": "",
        "WACC": "",
        "Risk Factor": "",
        "Revenue Growth": "",
        "Quality of Management Team": "",
        "Quality of Staff": ""
    }

    for ent in doc.ents:
        if ent.label_ == "ORG" and not data["Company Name"]:
            data["Company Name"] = ent.text
        elif ent.label_ == "GPE" and not data["Location"]:
            data["Location"] = ent.text
        elif ent.label_ == "DATE" and not data["Date"]:
            data["Date"] = ent.text

    financial_patterns = {
        "Revenue": r"Revenue[\s:]*\$?([\d,.]+)",
        "EBITDA": r"EBITDA[\s:]*\$?([\d,.]+)",
        "Gross Revenues": r"Gross Revenues[\s:]*\$?([\d,.]+)",
        "Owner salary": r"Owner salary[\s:]*\$?([\d,.]+)",
        "COGS": r"COGS[\s:]*\$?([\d,.]+)",
        "Gross Profit": r"Gross Profit[\s:]*\$?([\d,.]+)",
        "Total Operating Expenses": r"Total Operating Expenses[\s:]*\$?([\d,.]+)",
        "Operating Income": r"Operating Income[\s:]*\$?([\d,.]+)",
        "Net Pre-Tax Income": r"Net Pre-Tax Income[\s:]*\$?([\d,.]+)",
        "Adjusted EBITDA": r"Adjusted EBITDA[\s:]*\$?([\d,.]+)"
    }

    for key, pattern in financial_patterns.items():
        match = re.search(pattern, text)
        if match:
            data[key] = match.group(1)

    return data

def save_to_excel(data, filename="valuation_template.xlsx"):
    """Append extracted financial data to an Excel file or create a new file if it doesn't exist."""
    df = pd.DataFrame([data])

    try:
        book = load_workbook(filename)
        sheet = book.active

        with pd.ExcelWriter(filename, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            df.to_excel(writer, index=False, header=False, startrow=sheet.max_row)
    except FileNotFoundError:
        df.to_excel(filename, index=False)

    return filename

# Streamlit UI
st.title("Financial Data Extraction from PDF")

uploaded_file = st.file_uploader("Upload a PDF document", type=["pdf"])

if uploaded_file is not None:
    text = extract_text_from_pdf(uploaded_file)
    financial_data = extract_data(text)
    excel_file = save_to_excel(financial_data)

    st.success("Extraction completed!")
    st.download_button("Download Excel File", excel_file, file_name="valuation_template.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
